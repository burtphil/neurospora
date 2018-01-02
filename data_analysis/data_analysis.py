# -*- coding: utf-8 -*-


"""
Created on Tue Mar 14 13:10:38 2017

@author: Win7
"""

import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import os
import scipy.interpolate
from scipy import optimize
import matplotlib.mlab as mlab
from scipy import stats

#from mpl_toolkits.mplot3d.axes3d import Axes3D
#%%for LaTeX
'''
plt.rc('text', usetex=True)
plt.rc('font', family='serif')

plt.rcParams['xtick.labelsize'] = 26
plt.rcParams['ytick.labelsize'] = 26
'''
#%%Sometimes annoying
os.chdir(r'C:\Users\Philipp\Desktop\uni_remote\neurospora\all_raw_data')
z = os.getcwd()
q = os.listdir(z)
#%%
def Excel_to_array(name):
    chart = openpyxl.load_workbook(name, data_only=True)
    names = chart.get_sheet_names()
    sheet = chart.get_sheet_by_name(names[0])
    
    row_ind = int((sheet.max_column-1)/3)
    
    strains = []
    
    t = np.arange(sheet.max_row-1)
    arr = np.ones((4,len(t)))
    arr2 = np.copy(arr)
    arr3 = np.copy(arr)
    arr4 = np.copy(arr)
    arr5 = np.copy(arr)
    values = [arr, arr2, arr3, arr4, arr5]
        
    
    for i in range(row_ind):
        for j in range(len(t)):
            values[i][0,j] = sheet.cell(row=j+2, column=1).value
            values[i][1,j] = sheet.cell(row=j+2, column=1+i*3+1).value
            values[i][2,j] = sheet.cell(row=j+2, column=2+i*3+1).value
            values[i][3,j] = sheet.cell(row=j+2, column=3+i*3+1).value
        
        strain = sheet.cell(row=1, column=1+i*3+1).value
        strain = strain.split(' ')
        strain = strain[1]
        strains.append(strain)
        
        #x-offset
        values[i][0] -= values[i][0][0]
    
    T = int(name[14:16])
    warm = int(name[16:19])
    
    return (values, strains,T, warm)    
#%%
def Plotting(t, v1, v2, v3, strain, T, warm):
    To = T
    #t = (t/60)*(360./T) #in Grad
    #T *= (360./T)
    #midnight = ((T*(1-warm/100.))/2)
    #t -= T + midnight
            
    t = t/60. #in h
    
    title = str(strain[:-2]) + ', T = ' + str(To) + ', warm = ' + str(warm/100.)
    
    plt.figure(figsize=(10,6))
    plt.plot(t, v1, '.')
    plt.plot(t, v2, '.')
    plt.plot(t, v3, '.')
    plt.xticks(np.arange(0, len(t), T))
    plt.xlabel('t (h)',fontsize=18, labelpad=10)
    plt.ylabel('relative pixel density',fontsize=18, labelpad=5)
    plt.title(title, fontsize=14)
    
    ##Balken: dunkel = Nacht/kalt
    for i in range(int(t[-1]/T)+1):
        plt.axvspan(i*T, (1-warm/100)*T+i*T, facecolor='grey', alpha=0.2)
    
    datei = 'plots/' + str(strain[:-2]) + '_T' + str(To) + '_' + str(warm/100.) + '.png'
    plt.savefig(datei, bbox_inches='tight')
    
    #plt.xlim(3*T, 12*T)
    #plt.ylim(-200, 400)
    #plt.show()
#%%Plottet Mittelwert und Staabw aus Mittel_Stabw(values) -->not in use
'''
def Plotting2(t, v1, v2, v3, strain, T, warm):
    plt.figure(figsize=(10,6))
    plt.plot(t, v1, '.')
    plt.plot(t, v2, color = '0.9')
    plt.plot(t, v3, color = '0.9')    
    plt.fill_between(t, v2, v3, color = '0.9')
    plt.xlabel('t',fontsize=14)
    plt.ylabel('Amplitude',fontsize=14)
    plt.title(strain,fontsize=14)
    Zeitgeber(T, warm)
    plt.xlim(0,13000)
'''
#%%
def Mittel_Stabw(values):
    mean = np.copy(values)
    s_2 = np.ones(len(mean[1]))
    
    mean[1] = (values[1]+values[2]+values[3])/3
    s_2 = ((values[1]-mean[1])**2 + (values[2]-mean[1])**2 + (values[3]-mean[1])**2)/2
    mean[2] = mean[1] + np.sqrt(s_2)
    mean[3] = mean[1] -np.sqrt(s_2)
    
    return mean
#%%
def Fourier_Plot(t, v1, v2, v3, strain, T, warm):
    values = [v1, v2, v3]
    title = str(strain[:-2]) + ', T = ' + str(T) + ', warm = ' + str(warm/100.)
    plt.figure(figsize=(10,6))
    
    for i, v in enumerate(values):
        #window = np.hanning(len(v))
        #window = np.hamming(len(v))
        window = np.blackman(len(v))
        #window = 1 #'Rechteck'
        
        zeropad = 0
        
        fourier = np.fft.rfft(v*window, len(v)+zeropad)#already scaled by 1/n!
        freqs = np.fft.rfftfreq(len(v)+zeropad, 10)  #anzahl messwerte, abstand messwerte
        taus = np.ones(len(freqs))
        for i, v in enumerate(freqs):
            if v!=0: taus[i] = 1./v
            else: taus[i] = 0
        taus /= 60.
        
        peak = np.argmax(np.absolute(fourier)**2)
        tau = np.round(taus[peak], 2)
        if tau<0: tau*=-1
        
        plt.plot(taus, np.absolute(fourier)**2, 'o-', label = tau) #power spectrum
        
    plt.title(title, fontsize=14)
    plt.legend()
    plt.ylim(0,)
    plt.xlabel('entrained period (h)',fontsize=18, labelpad=10)
    plt.ylabel('power',fontsize=18, labelpad=5)
    
    datei = 'fourier/' + str(strain[:-2]) + '_T' + str(T) + '_' + str(warm/100.) + '_power.png'
    plt.savefig(datei, bbox_inches='tight')
    
    #plt.show()
#%%
def fitfunc(p, x):
    return p[0]*np.cos(p[1]*x+p[2])

def errfunc(p, x, y):
    return fitfunc(p, x) - y

def fitparabola(p,x):
    return p[0]*x**2 + p[1]*x + p[2]

def errparabola(p, x, y):
    return fitparabola(p, x) - y

def fitline(p,x):
    return p[0]*x + p[1]

def errline(p, x, y):
    return fitline(p, x) - y
#%%
def Tau_Phi(t, va1 , va2, va3, T, warm):
    values = [va1, va2, va3]
    taus = []
    phis = []
    To = T
    T = T*(360./T)
    warm = warm/100.
    
    t = (t/60)*(360./To) #in Grad
    midnight = ((T*(1-warm))/2)
    t -= (T + midnight)
    
    fourier = np.fft.fft(va1)
    freqs = np.fft.fftfreq(len(va1), 10)  #anzahl messwerte, abstand messwerte
    peak = np.argmax(np.absolute(fourier)**2)
    tau = 1/-freqs[peak] #in Minuten
    tau = tau/60. * (360./To) #in °
    omega = 2*np.pi/tau
    
    for i, v in enumerate(values):
        
        p0 = [400, omega, 0]
        fit, success = optimize.leastsq(errfunc, p0, args=(t, v)) 
        tau = 2*np.pi/fit[1]
        tauh = tau*(To/360.)
        zero = ((np.pi/2-fit[2])/fit[1])
        if fit[1]<0: tau *= -1; tauh *= -1; zero += (3*tau/2.)
        phi = zero - 0
        phih = phi*(To/360.) #Zeitgeber-Tag in h
        phiex = phih*(24./To) #exT in h
        psi = phiex*(360./24)
        
        taus.append(tauh) #in h
        phis.append(psi) #in° -->phi eventuell besser, da Connie nicht nachvollziehbar
        
        #plt.plot(t, v, '.')
        #plt.plot(t, fitfunc(fit,t))
        #plt.axvspan(0-(T*(1-warm)/2), 0+(T*(1-warm)/2), facecolor='blue', alpha=0.3)
        #plt.axvspan(0-(T*(1-warm)/2)-T, 0+(T*(1-warm)/2)-T, facecolor='blue', alpha=0.3)
        #plt.axvspan(0-(T*(1-warm)/2)+T, 0+(T*(1-warm)/2)+T, facecolor='blue', alpha=0.3)
        #plt.plot(zero, 0, 'ok')
        #plt.plot(0, 0, 'ob')
        #plt.title(tauh)
        #plt.xlim(-360, 360)
        #plt.show()
    
    taus = np.asarray(taus)
    taus = taus[~np.isnan(taus)]
        
    tau_ = np.sum(taus)/len(taus)
    phi_ = sum(phis)/len(phis)
    
    return (tau_ , phi_)

#%%
def Interpolate_Max(v):
    maxi = np.amax(v)
    if list(v).count(maxi)!=1:
        y = np.array([v[(np.argwhere(v==maxi)-1)[0,0]], maxi, maxi, v[(np.argwhere(v==maxi)+1)[1,0]]])
        x = np.array([1,2,3,4])
    else: 
        y = np.array([v[np.argwhere(v==maxi)-2], v[np.argwhere(v==maxi)-1], maxi, v[np.argwhere(v==maxi)+1], v[np.argwhere(v==maxi)+2]])
        x = np.array([1,2,3,4,5])
    fc = scipy.interpolate.interp1d(x, y,kind='cubic')
    truemax = np.amax(fc(np.arange(np.amin(x),np.amax(x),0.2)))
    return truemax   
#%%
def Amplitude(t, v1, v2, v3):
    values = [v1, v2, v3]
    maxima = []
    minima = []
    for i, v in enumerate(values):
        indices = [0, len(v)-1]
        zero_crossings = np.where(np.diff(np.signbit(v)))[0]
        indices += np.ndarray.tolist(zero_crossings)
        indices = sorted(indices)

        lmax = []
        lmin = []
        for j in range(len(indices)-2):#das letzte Intervall bleibt unberücksichtigt
            if v[indices[j]]>0:
                mini = np.amin(v[indices[j]:indices[j+1]])
                
                #Regression, die Werte sind zu niedrig
                #p0 = [1, 1, 0]
                #fit, success = optimize.leastsq(errparabola, p0, args=(t[indices[j]:indices[j+1]], v[indices[j]:indices[j+1]]))
                #mini = -(fit[1]**2/(4*fit[0]))+fit[2]
                
                lmin.append(mini)
            elif v[indices[j]]<=0:
                maxi = np.amax(v[indices[j]:indices[j+1]])
                if maxi>30:
                    maxi = Interpolate_Max(v[indices[j]:indices[j+1]])
                    lmax.append(maxi)
        
        lmax = np.asarray(lmax)
        lmin = np.asarray(lmin)
        
        lmax = lmax[lmax>125]
        lmin = lmin[lmin<-30]
        
        lmax = lmax[~np.isnan(lmax)]
        lmin = lmin[~np.isnan(lmin)]
        
        maxima.append(np.sum(lmax)/len(lmax))
        minima.append(np.sum(lmin)/len(lmin))
        
    maxima = np.asarray(maxima)
    minima = np.asarray(minima)
    
    maxima = maxima[~np.isnan(maxima)]
    minima = minima[~np.isnan(minima)]
        
    return (maxima, minima)
    
#indices ist korrekt(teilweise ungenau, aber okay)
#%%    
print(q)
#%%
current = Excel_to_array('con-data for t16 60-40_ b.xlsx')
print(current[1])
#%%
a = 2
Plotting(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3], current[1][a], current[2], current[3])
Fourier_Plot(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3], current[1][a], current[2], current[3])
"""
#%%alle plots
tables=q[6:-8]
for i in range(len(tables)):
    current = Excel_to_array(tables[i])
    for k in range(len(current[1])):
        a = k
        Plotting(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3], current[1][a], current[2], current[3])
        Fourier_Plot(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3], current[1][a], current[2], current[3])
#%%
#mean = Mittel_Stabw(current[0][a])
#%% get Tau and Phi
res = Tau_Phi(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3], current[2], current[3])
print(res[0], res[1])
#%%get amplitude
amps = Amplitude(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3])
print((amps[0]-amps[1])/2., amps[0], amps[1])
#%%

#tabellen = ['con-data for t16 16-84_ b.xlsx', 'con-data for t16 25-75_ b.xlsx', 'con-data for t16 33-67_ b.xlsx', 'con-data for t16 40-60_ b.xlsx', 'con-data for t16 50-50_ b.xlsx', 'con-data for t16 60-40_ b.xlsx', 'con-data for t16 67-33_ b.xlsx', 'con-data for t16 75-25_ b.xlsx', 'con-data for t16 84-16_ b.xlsx', 
#            'con-data for t22 16-84_ b.xlsx', 'con-data for t22 25-75_ b.xlsx', 'con-data for t22 33-67_ b.xlsx', 'con-data for t22 40-60_ b.xlsx', 'con-data for t22 50-50_ b.xlsx', 'con-data for t22 60-40_ b.xlsx', 'con-data for t22 67-33_ b.xlsx', 'con-data for t22 75-25_ b.xlsx', 'con-data for t22 84-16_ b.xlsx',
#            'con-data for t26 16-84_ b.xlsx', 'con-data for t26 25-75_ b.xlsx', 'con-data for t26 33-67_ b.xlsx', 'con-data for t26 40-60_ b.xlsx', 'con-data for t26 50-50_ b.xlsx', 'con-data for t26 60-40_ b.xlsx', 'con-data for t26 67-33_ b.xlsx', 'con-data for t26 75-25_ b.xlsx', 'con-data for t26 84-16_ b.xlsx']

tabellen = ['con-data for t16 16-84_ a.xlsx', 'con-data for t16 25-75_ a.xlsx', 'con-data for t16 33-67_ a.xlsx', 'con-data for t16 40-60_ a.xlsx', 'con-data for t16 50-50_ a.xlsx', 'con-data for t16 60-40_ a.xlsx', 'con-data for t16 67-33_ a.xlsx', 'con-data for t16 75-25_ a.xlsx', 'con-data for t16 84-16_ a.xlsx', 
            'con-data for t22 16-84_ a.xlsx', 'con-data for t22 25-75_ a.xlsx', 'con-data for t22 33-67_ a.xlsx', 'con-data for t22 40-60_ a.xlsx', 'con-data for t22 50-50_ a.xlsx', 'con-data for t22 60-40_ a.xlsx', 'con-data for t22 67-33_ a.xlsx', 'con-data for t22 75-25_ a.xlsx', 'con-data for t22 84-16_ a.xlsx',
           'con-data for t26 16-84_ a.xlsx', 'con-data for t26 25-75_ a.xlsx', 'con-data for t26 33-67_ a.xlsx', 'con-data for t26 40-60_ a.xlsx', 'con-data for t26 50-50_ a.xlsx', 'con-data for t26 60-40_ a.xlsx', 'con-data for t26 67-33_ a.xlsx', 'con-data for t26 75-25_ a.xlsx', 'con-data for t26 84-16_ a.xlsx']
            
#obere tabellen a = 1,2; untere a=2            
a = 2
ziel = np.ones([len(tabellen),5])
for i,v in enumerate(tabellen):
    current = Excel_to_array(v)
    
    #Plotting(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3], current[1][a], current[2], current[3])
    
    amps = Amplitude(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3])
    amp = (np.sum(amps[0]-amps[1])/2.)/len(amps[0])
    
    res = Tau_Phi(current[0][a][0], current[0][a][1], current[0][a][2], current[0][a][3], current[2], current[3])
    ziel[i][0]=22. #tau
    ziel[i][3]=res[0] #T entrained
    ziel[i][1]=current[2] #T set-up
    ziel[i][2]=current[3]/100. #warm
    ziel[i][4]=amp #amplitude

#np.savetxt('tau_22_amps.txt', ziel, delimiter="\t")
#%%
chart2 = openpyxl.load_workbook('Amplituden6.xlsx', data_only=True)
sheet2 = chart2.get_sheet_by_name('sheet1')
array = np.ones([sheet2.max_column, sheet2.max_row-1])
for i in range(len(array[0])):
            array[0][i] = sheet2.cell(row=i+2, column=1).value
            array[1][i] = sheet2.cell(row=i+2, column=2).value
            array[2][i] = sheet2.cell(row=i+2, column=3).value
            array[3][i] = sheet2.cell(row=i+2, column=4).value
            array[4][i] = sheet2.cell(row=i+2, column=5).value
            array[5][i] = sheet2.cell(row=i+2, column=6).value
            array[6][i] = sheet2.cell(row=i+2, column=7).value
                 
#%%K vs psi -->sortiert nach tau

for j in range(3):
    plt.figure(figsize=(10,6))
    for i in range(3):
        za = np.copy(5/array[4][0+j*27+i*9:9+j*27+i*9]*100)
        zb = np.copy(array[5][0+j*27+i*9:9+j*27+i*9])
        zc = np.copy(array[6][0+j*27+i*9:9+j*27+i*9])
        zd = np.copy(array[2][0+j*27+i*9:9+j*27+i*9])
        #za *= zd#thermoperiod
        za = za[za.argsort()] #für o-
        zb = zb[za.argsort()]
        za = za[zc<2]#keine nicht entrainten
        zb = zb[zc<2]
        for k,v in enumerate(zb):#da zyklisch
            if v>15: zb[k]=zb[k]-24
        #p0 = [1, 1]
        #if len(za)>0:fit, success = optimize.leastsq(errline, p0, args=(za, zb))
        label = 'T = '+str(array[1][i*9])
        title = 'tau = '+str(array[0][j*27])
        plt.plot(za, zb, 'o-', label=label)
        #if len(za)>0:plt.plot(za, fitline(fit, za), 'k', lw=0.5)
    plt.title(title)
    plt.xlabel('Z/A',fontsize=14)
    plt.ylabel('psi',fontsize=14)
    plt.legend(loc=1)
#%%K vs psi -->sortiert nach T
for j in range(3):
    plt.figure(figsize=(10,6))
    for i in range(3):
        za = np.copy(5/array[4][0+i*27+j*9:9+i*27+j*9]*100)
        zb = np.copy(array[5][0+i*27+j*9:9+i*27+j*9])
        zc = np.copy(array[6][0+i*27+j*9:9+i*27+j*9])
        zd = np.copy(array[2][0+i*27+j*9:9+i*27+j*9])
        #za *= zd#thermoperiod
        za = za[za.argsort()] #für o-
        zb = zb[za.argsort()]
        za = za[zc<2]#keine nicht entrainten
        zb = zb[zc<2]
        for k,v in enumerate(zb):#da zyklisch
            if v>15: zb[k]=zb[k]-24
        #p0 = [1, 1]
       # if len(za)>0:fit, success = optimize.leastsq(errline, p0, args=(za, zb))
        label = 'T = '+str(array[1][j*9])
        title = 'tau = '+str(array[0][i*27])
        plt.plot(za, zb, 'o-', label=title)
        #if len(za)>0:plt.plot(za, fitline(fit, za), 'k', lw=0.5)
    plt.title(label)
    plt.xlabel('Z/A',fontsize=14)
    plt.ylabel('psi',fontsize=14)
    plt.legend(loc=1)
#%%therm vs psi-->sortiert nach tau, Remi/Connie waagerecht
for j in range(3):
    plt.figure(figsize=(10,6))
    for i in range(3):
        zb = np.copy(array[5][0+j*27+i*9:9+j*27+i*9])#psi
        zc = np.copy(array[6][0+j*27+i*9:9+j*27+i*9])#entrainment-check
        zd = np.copy(array[2][0+j*27+i*9:9+j*27+i*9])#thermoperiod
        zd = zd[zd.argsort()] #für o-
        zb = zb[zd.argsort()]
        zd = zd[zc<2]#keine nicht entrainten
        zb = zb[zc<2]
        for k,v in enumerate(zb):#da zyklisch
            if v>15: zb[k]=zb[k]-24
        #p0 = [1, 1]
        #if len(zd)>0:fit, success = optimize.leastsq(errline, p0, args=(zd, zb))
        label = 'T = '+str(array[1][i*9])
        title = 'tau = '+str(array[0][j*27])
        plt.plot(zd, zb, 'o-', label=label)
    #if len(zd)>0:plt.plot(zd, fitline(fit, zd), 'k', lw=0.5)
    plt.title(title)
    plt.xlabel('thermoperiod',fontsize=14)
    plt.ylabel('psi (h)',fontsize=14)
    plt.legend(loc=1)
#%%therm vs psi-->sortiert nach T
for j in range(3):
    plt.figure(figsize=(10,6))
    for i in range(3):
        zb = np.copy(array[5][0+i*27+j*9:9+i*27+j*9])#psi
        zc = np.copy(array[6][0+i*27+j*9:9+i*27+j*9])#entrainment-check
        zd = np.copy(array[2][0+i*27+j*9:9+i*27+j*9])#thermoperiod
        zd = zd[zd.argsort()] #für o-
        zb = zb[zd.argsort()]
        zd = zd[zc<2]#keine nicht entrainten
        zb = zb[zc<2]
        for k,v in enumerate(zb):#da zyklisch
            if v>15: zb[k]=zb[k]-24
        #p0 = [1, 1]
        #if len(zd)>0:fit, success = optimize.leastsq(errline, p0, args=(zd, zb))
        label = 'T = '+str(array[1][j*9])
        title = 'tau = '+str(array[0][i*27])
        plt.plot(zd, zb, 'o-', label=title)
    #if len(zd)>0:plt.plot(zd, fitline(fit, zd), 'k', lw=0.5)
    plt.title(label)
    plt.xlabel('thermoperiod',fontsize=14)
    plt.ylabel('psi (h)',fontsize=14)
    plt.legend(loc=1)
#%%therm vs A-->sortiert nach tau -->Resonanzen
for j in range(3):
    plt.figure(figsize=(10,6))
    for i in range(3):
        zb = np.copy(array[4][0+j*27+i*9:9+j*27+i*9])
        zc = np.copy(array[6][0+j*27+i*9:9+j*27+i*9])
        zd = np.copy(array[2][0+j*27+i*9:9+j*27+i*9])
        zd = zd[zd.argsort()] #für o-
        zb = zb[zd.argsort()]
        zd = zd[zc<2]#keine nicht entrainten
        zb = zb[zc<2]
        for k,v in enumerate(zb):#da zyklisch
            if v>15: zb[k]=zb[k]-24
        p0 = [1, 1]
        if len(zd)>0:fit, success = optimize.leastsq(errline, p0, args=(zd, zb))
        label = 'T = '+str(array[1][i*9])
        title = 'tau = '+str(array[0][j*27])
        plt.plot(zd, zb, 'o-', label=label)
        stichp1 = np.append(zb[:2], zb[5:])
        stichp2 = zb[3:6]
        pval = stats.ttest_ind(stichp1, stichp2)
        #print(pval)
        #if len(zd)>0:plt.plot(zd, fitline(fit, zd), 'k', lw=0.5);print(fit[0])
    plt.title(title)
    plt.xlabel('thermoperiod',fontsize=14)
    plt.ylabel('A',fontsize=14)
    plt.legend(loc=4)
#%%therm vs A -->sortiert nach T
for j in range(3):
    plt.figure(figsize=(10,6))
    for i in range(3):
        zb = np.copy(array[4][0+i*27+j*9:9+i*27+j*9])
        zc = np.copy(array[6][0+i*27+j*9:9+i*27+j*9])
        zd = np.copy(array[2][0+i*27+j*9:9+i*27+j*9])
        zd = zd[zd.argsort()] #für o-
        zb = zb[zd.argsort()]
        zd = zd[zc<2]#keine nicht entrainten
        zb = zb[zc<2]
        for k,v in enumerate(zb):#da zyklisch
            if v>15: zb[k]=zb[k]-24
        p0 = [1, 1]
        if len(zd)>0:fit, success = optimize.leastsq(errline, p0, args=(zd, zb))
        label = 'T = '+str(array[1][j*9])
        title = 'tau = '+str(array[0][i*27])
        plt.plot(zd, zb, 'o-', label=title)
        stichp1 = np.append(zb[:2], zb[5:])
        stichp2 = zb[3:6]
        pval = stats.ttest_ind(stichp1, stichp2)
        #print(pval)
        #if len(zd)>0:plt.plot(zd, fitline(fit, zd), 'k', lw=0.5);print(fit[0])
    plt.title(label)
    plt.xlabel('thermoperiod',fontsize=14)
    plt.ylabel('A',fontsize=14)
    plt.legend(loc=4)
#%%T vs psi für jeden strain
fits2 =[]
therms2 = []
for l in range(2):#für tau=27 zu wenige werte
    plt.figure(figsize=(10,6))
    fits = []
    therms = []
    for i in range(9):
        Ts = []
        psis = []
        for j in range(3):
            if array[6][i+9*j+27*l]<2:
                psis.append(array[5][i+9*j+27*l])
                Ts.append(array[1][i+9*j+27*l])
                for k,v in enumerate(psis):#da zyklisch
                    if v>15: psis[k]=psis[k]-24
        if len(psis)==3:
            p0 = [-1, 1]
            Ts = np.asarray(Ts)
            psis = np.asarray(psis)            
            fit, success, info, mesq, ier = optimize.leastsq(errline, p0, args=(Ts,psis),full_output=True) #info: dictionary mit residuals
            ss_err=(info['fvec']**2).sum()#residual sum of squares
            ss_tot=((psis-psis.mean())**2).sum()#total sum of squares
            rsquared=1-(ss_err/ss_tot)
            fits.append(fit[0])
            therms.append(array[2][i])
            plt.plot(Ts, psis, 'o', label=str(array[2][i])+';  R^2 = '+str(np.round(rsquared,2)))
            plt.plot(Ts, fitline(fit, Ts), 'k', lw=0.5)
    plt.legend()
    plt.xlabel('T (h)',fontsize=14)
    plt.ylabel('psi (h)',fontsize=14)
    plt.title('tau = '+str(array[0][0+l*27]),fontsize=14)
    fits2.append(fits)
    therms2.append(therms)

plt.figure(figsize=(10,6))
plt.plot(therms2[0], fits2[0], 'o', label='tau = 16.5')
plt.plot(therms2[1], fits2[1], 'o', label='tau = 22')
plt.xlabel('thermoperiod',fontsize=14)
plt.ylabel('psi/T',fontsize=14)
plt.legend()   
#%%tau vs psi für jedes T
fits2 =[]
therms2 = []
for l in range(2):
    l += 1 #nicht genug entrainment für T = 16
    plt.figure(figsize=(10,6))
    fits = []
    therms = []
    for i in range(9):
        Ts = []
        psis = []
        for j in range(3):
            if array[6][i+27*j+9*l]<2:
                psis.append(array[5][j*27+i+9*l])
                Ts.append(array[0][j*27+i+9*l])
                label = str(array[2][j*27+i+9*l])
            for k,v in enumerate(psis):#da zyklisch
                if v>15: psis[k]=psis[k]-24
        if len(psis)==3:
            p0 = [-1, 1]
            Ts = np.asarray(Ts)
            psis = np.asarray(psis)
            fit, success, info, mesq, ier = optimize.leastsq(errline, p0, args=(Ts,psis), full_output=True)
            ss_err=(info['fvec']**2).sum()
            ss_tot=((psis-psis.mean())**2).sum()
            rsquared=1-(ss_err/ss_tot)
            fits.append(fit[0])
            therms.append(array[2][j*27+i+9*l])
            plt.plot(Ts, psis, 'o', label=label+';  R^2 = '+str(np.round(rsquared,2)))
            plt.plot(Ts, fitline(fit, Ts), 'k', lw=0.5)
    plt.legend()
    plt.xlabel('tau (h)',fontsize=14)
    plt.ylabel('psi (h)',fontsize=14)
    plt.title('T = '+str(array[1][0+l*9]),fontsize=14)
    fits2.append(fits)
    therms2.append(therms)
    
plt.figure(figsize=(10,6))
plt.plot(therms2[0], fits2[0], 'o', label='T = 22')
plt.plot(therms2[1], fits2[1], 'o', label='T = 26')
plt.xlabel('thermoperiod',fontsize=14)
plt.ylabel('psi/tau',fontsize=14)
plt.legend()

#%%detuning vs psi
plt.figure(figsize=(10,7))
fits2 =[]
therms2 = []
colors = ['r', 'b', 'g', 'r', 'g', 'g', 'r', 'r', 'g']
markers = ['o', 'o', 'o', 'o', 'o', 'o', 'o', 'o', 'o']
for l in range(3):
    fits = []
    therms = []
    for i in range(9):
        Ts = []
        psis = []
        for j in range(3):
            if array[6][i+9*j+27*l]<2:
                psis.append(array[5][i+9*j+27*l])
                Ts.append(array[0][i+9*j+27*l]-array[1][i+9*j+27*l])
                for k,v in enumerate(psis):#da zyklisch
                    if v>15: psis[k]=psis[k]-24
                #label = str(array[2][i])+', '+str(array[0][i+9*j+27*l])
                label = r'$\kappa$ = '+str(array[2][i])
        if len(psis)==3:
            p0 = [-1, 1]
            Ts = np.asarray(Ts)
            psis = np.asarray(psis)
            fit, success, info, mesq, ier = optimize.leastsq(errline, p0, args=(Ts,psis), full_output=True)
            ss_err=(info['fvec']**2).sum()
            ss_tot=((psis-psis.mean())**2).sum()
            rsquared=1-(ss_err/ss_tot)
            fits.append(fit[0])
            therms.append(array[2][i])
            if array[0][i+9*j+27*l]<20: plt.plot(Ts, fitline(fit, Ts), color='b', alpha=i*0.1+0.1);
            #elif array[0][i+9*j+27*l]>20 and array[0][i+9*j+27*l]<25: plt.plot(Ts, fitline(fit, Ts), color='g', alpha=i*0.1);
            
        if array[0][i+9*j+27*l]<20: plt.plot(Ts, psis, 'o', markersize=12, linestyle='None',  color = 'b', alpha=i*0.1+0.1, label=label);
        #elif array[0][i+9*j+27*l]>20 and array[0][i+9*j+27*l]<25: plt.plot(Ts, psis, '^', markersize=12, linestyle='None', color='g', alpha=i*0.1);
        #elif array[0][i+9*j+27*l]>25: plt.plot(Ts, psis, 's', markersize=12, linestyle='None',  color = 'r', alpha=i*0.1+0.1);
                
    plt.legend(fontsize=14)
    plt.xlabel(r'$\tau - T$ (h)',fontsize=26, labelpad=20)  
    plt.ylabel(r'$\psi$  (h)',fontsize=26, labelpad=20)
    plt.xticks(np.arange(-10., 1, 2))
    fits2.append(fits)
    therms2.append(therms)
#%%
plt.figure(figsize=(10,7))    
plt.plot(therms2[0], fits2[0], 'ob', markersize=12, label='tau = 16.5')
#plt.plot(therms2[1], fits2[1], '<g', label='tau = 22')
plt.xlabel(r'$\kappa$', fontsize=26, labelpad=20)
plt.ylabel(r'$slope \;(\frac{\psi}{\tau-T})$', fontsize=20, labelpad=20)
#plt.legend()
#%%T vs A für jeden strain

fits2 =[]
therms2 = []
for l in range(2):#für tau=27 zu wenige werte
    plt.figure(figsize=(10,6))
    fits = []
    therms = []
    for i in range(9):
        Ts = []
        psis = []
        for j in range(3):
            if array[6][i+9*j+27*l]<2:
                psis.append(array[4][i+9*j+27*l])
                Ts.append(array[1][i+9*j+27*l])
                for k,v in enumerate(psis):#da zyklisch
                    if v>15: psis[k]=psis[k]
        if len(psis)==3:
            p0 = [-1, 1]
            Ts = np.asarray(Ts)
            psis = np.asarray(psis)
            fit, success, info, mesq, ier = optimize.leastsq(errline, p0, args=(Ts,psis), full_output=True)
            ss_err=(info['fvec']**2).sum()
            ss_tot=((psis-psis.mean())**2).sum()
            rsquared=1-(ss_err/ss_tot)
            fits.append(fit[0])
            therms.append(array[2][i])
            plt.plot(Ts, psis, 'o', label=str(array[2][i])+';  R^2 = '+str(np.round(rsquared,2)))
            plt.plot(Ts, fitline(fit, Ts), 'k', lw=0.5)
    plt.legend()
    plt.xlabel('T (h)',fontsize=14)
    plt.ylabel('A',fontsize=14)
    plt.title('tau = '+str(array[0][0+l*27]),fontsize=14)
    fits2.append(fits)
    therms2.append(therms)
    
plt.figure(figsize=(10,6))
plt.plot(therms2[0], fits2[0], 'o', label='tau = 16.5')
plt.plot(therms2[1], fits2[1], 'o', label='tau = 22')
plt.xlabel('thermoperiod',fontsize=14)
plt.ylabel('A/T',fontsize=14)
plt.legend()
#%%tau vs A für jedes T
fits2 =[]
therms2 = []
for l in range(2):
    l += 1 #nicht genug entrainment für T = 16
    plt.figure(figsize=(10,6))
    fits = []
    therms = []
    for i in range(9):
        Ts = []
        psis = []
        for j in range(3):
            if array[6][i+27*j+9*l]<2:
                psis.append(array[4][j*27+i+9*l])
                Ts.append(array[0][j*27+i+9*l])
                label = str(array[2][j*27+i+9*l])
            for k,v in enumerate(psis):#da zyklisch
                if v>15: psis[k]=psis[k]
        if len(psis)==3:
            p0 = [-1, 1]
            Ts = np.asarray(Ts)
            psis = np.asarray(psis)
            fit, success, info, mesq, ier = optimize.leastsq(errline, p0, args=(Ts,psis), full_output=True)
            ss_err=(info['fvec']**2).sum()
            ss_tot=((psis-psis.mean())**2).sum()
            rsquared=1-(ss_err/ss_tot)
            fits.append(fit[0])
            therms.append(array[2][i])
            plt.plot(Ts, psis, 'o', label=label+';  R^2 = '+str(np.round(rsquared,2)))
            plt.plot(Ts, fitline(fit, Ts), 'k', lw=0.5)
    plt.legend()
    plt.xlabel('tau (h)',fontsize=14)
    plt.ylabel('A',fontsize=14)
    plt.title('T = '+str(array[1][0+l*9]),fontsize=14)
    fits2.append(fits)
    therms2.append(therms)
    
plt.figure(figsize=(10,6))
plt.plot(therms2[0], fits2[0], 'o', label='T = 22')
plt.plot(therms2[1], fits2[1], 'o', label='T = 26')
plt.xlabel('thermoperiod',fontsize=14)
plt.ylabel('A/tau',fontsize=14)
plt.legend()

#%%detuning vs A
plt.figure(figsize=(10,6))
fits2 =[]
therms2 = []
for l in range(3):
    fits = []
    therms = []
    stichp1 = []
    stichp2 = []
    for i in range(9):
        Ts = []
        psis = []
        for j in range(3):
            if array[6][i+9*j+27*l]<2:
                psis.append(array[4][i+9*j+27*l])
                Ts.append(array[0][i+9*j+27*l]-array[1][i+9*j+27*l])
                for k,v in enumerate(psis):#da zyklisch
                    if v>15: psis[k]=psis[k]
                label = str(array[2][i])+', '+str(array[0][i+9*j+27*l])
        if len(psis)==3:
            p0 = [-1, 1]
            Ts = np.asarray(Ts)
            psis = np.asarray(psis)
            fit, success, info, mesq, ier = optimize.leastsq(errline, p0, args=(Ts,psis), full_output=True)
            ss_err=(info['fvec']**2).sum()
            ss_tot=((psis-psis.mean())**2).sum()
            rsquared=1-(ss_err/ss_tot)
            fits.append(fit[0])
            therms.append(array[2][i])
            if array[0][i+9*j+27*l]<20:plt.plot(Ts, psis, 'o', color='b', alpha=i*0.1+0.2, label=label+';  R^2 = '+str(np.round(rsquared,2)))
            else: plt.plot(Ts, psis, '<', color='g', alpha=i*0.1+0.2, label=label+';  R^2 = '+str(np.round(rsquared,2)))
           # plt.plot(Ts, fitline(fit, Ts), 'k', lw=0.5)
            stichp1.append(psis[2])
            stichp1.append(psis[1])
            stichp2.append(psis[0])
        pval = stats.ttest_ind(np.asarray(stichp1), np.asarray(stichp2))
    print(pval) #only for tau = 16!!! [upper value]
    print(np.mean(stichp1), np.mean(stichp2))
    #plt.legend()
    plt.xlabel('tau - T (h)',fontsize=20, labelpad=20)
    plt.ylabel('A',fontsize=20, labelpad=20)
    fits2.append(fits)
    therms2.append(therms)

plt.figure(figsize=(10,6))    
plt.plot(therms2[0], fits2[0], 'o', label='tau = 16.5')
plt.plot(therms2[1], fits2[1], '<', label='tau = 22')
plt.xlabel('thermoperiod', fontsize=14)
plt.ylabel('A/(tau-T)', fontsize=14)
plt.legend()

#%%schwierig, die onion ist für unterschiedlche thermoperioden auch unterschiedlich schief
#--> eventuell mit deutlich! mehr Ts möglich
'''
for i in range(3):
    plt.figure(figsize=(10,6))
    for j in range(9):
        za = []
        zb = []
        za.append(5./array[4][j+0+27*i]*100)
        za.append(5./array[4][j+9+27*i]*100)
        za.append(5./array[4][j+18+27*i]*100)
        zb.append(array[0][j+0+27*i]-array[1][j+0+27*i])
        zb.append(array[0][j+9+27*i]-array[1][j+9+27*i])
        zb.append(array[0][j+18+27*i]-array[1][j+18+27*i])
        plt.plot(zb, za, 'o-', label=array[2][j])
    plt.legend()
    plt.show()

#%%
image = np.ones([9,3])
image[:,0] = array[6][0:9]
image[:,1] = array[6][9:18]
image[:,2] = array[6][18:27]
image = np.flipud(image)

image2 = np.ones([9,3])
image2[:,0] = array[6][27:36]
image2[:,1] = array[6][36:45]
image2[:,2] = array[6][45:54]
image2 = np.flipud(image2)

image3 = np.ones([9,3])
image3[:,0] = array[6][54:63]
image3[:,1] = array[6][63:72]
image3[:,2] = array[6][72:81]
image3 = np.flipud(image3)

plt.figure(figsize=(10,6))
plt.subplot(131)
plt.imshow(image)
plt.xticks(range(3), [16,22,26])
plt.yticks(range(9), np.flipud(array[2][0:9]))
plt.title('tau = 16,5h')
plt.subplot(132)
plt.title('tau = 22h')
plt.imshow(image2)
plt.xticks(range(3), [16,22,26])
plt.yticks(range(9), np.flipud(array[2][0:9]))
plt.subplot(133)
plt.imshow(image3)
plt.xticks(range(3), [16,22,26])
plt.yticks(range(9), np.flipud(array[2][0:9]))
plt.title('tau = 29h')
plt.show()
#%%

k = np.ones([2,9])
q = array[2][0:9].tolist()#0:9-->alle
q = np.round(q, 2)
#fig = plt.figure(figsize=plt.figaspect(0.5))
#ax = fig.add_subplot(111, projection='3d')
col = ['grey', 'g', 'r','c', 'm', 'orange','b', 'y', 'k']
for j, v in enumerate(q):
    for i in range(9):
        p = i*9
        k[0,i] = array[1][p+j] - array[0][p+j]
        k[1,i] = array[6][p+j]
    k = np.transpose(k)
    k = k[k[:,0].argsort()]
    k = np.transpose(k)
    
    X, Y = np.meshgrid(k[0], v)
    Z = k[1]    
    
    #ax.plot_wireframe(X, Y, Z, rstride=10, cstride=10, color=col[j])
    #ax.set_xlabel('T-tau')
   # ax.set_ylabel('photoperiod')
   # ax.set_zlabel('entrainment')
    plt.plot(k[0], k[1], 'x', label=v)
plt.xlabel('T - tau')
plt.ylabel('entrainment')
plt.legend(fancybox=True)
plt.show()
    #print(v, np.mean(k[1]), np.std(k[1]), np.mean(k[1][3:-3]), np.std(k[1][3:-3]))

#%%

k2 = np.ones([5,9])
qz = array[2][0:9].tolist()#0:9-->alle
qz = np.round(qz, 2)
plt.figure(figsize=(10,6))
for j, v in enumerate(qz):
    for i in range(9):
        p = i*9
        k2[0,i] = array[3][p+j] - array[0][p+j]
        k2[1,i] = array[3][p+j] - array[1][p+j]
        k2[2,i] = array[6][p+j]
        k2[3,i] = array[1][p+j] - array[0][p+j]
        k2[4,i] = array[3][p+j]
    #k2 = np.transpose(k)
    #k2 = k2[k2[:,0].argsort()]
    #k2 = np.transpose(k)
    
    #plt.plot(k2[0], k2[2], 'x', label=v) #T_ent-tau vs ent
    #plt.xlabel('t_ent - tau')
    plt.plot(k2[1], k2[2], 'x', label=v) #T_ent-T vs ent
    plt.xlabel('t_ent - T')
    plt.ylabel('estimated entrainment')
    plt.ylim(0,4.5)  
    
plt.legend(fancybox=True)
plt.show()

#%%
plt.rcParams['xtick.labelsize'] = 16
plt.rcParams['ytick.labelsize'] = 16
hiss = []
ex0 = []
ex1 = []
for j in range(3):
    his = [[],[]]
    for i, v in enumerate(array[6][0+j*27:27+j*27]):
        if v<2:
            if array[5][i+j*27]<15: psi = array[5][i+j*27]
            else: psi = (array[5][i+j*27]-24)
            his[0].append(psi)#psis
            his[1].append(array[4][i+j*27])#amplitudes
    hiss.append(his)
    ex0.append(max(his[0]))#für x-Achse psis
    ex0.append(min(his[0]))
    ex1.append(max(his[1]))#für x-Achse amplitudes
    ex1.append(min(his[1]))
    
x = np.arange(round(min(ex0)), round(max(ex0)), 1) #binbreite
hiss = np.asarray(hiss)

y1 = mlab.normpdf(x, np.mean(hiss[0][0]), np.std(hiss[0][0])) #16.5
y2 = mlab.normpdf(x, np.mean(hiss[1][0]), np.std(hiss[1][0])) #22
y3 = mlab.normpdf(x, np.mean(hiss[2][0]), np.std(hiss[2][0])) #29

f, ((ax1), (ax2), (ax3)) = plt.subplots(3,1, figsize=(12,8), sharex=True, sharey=True)
ax1.hist(hiss[0][0], x, edgecolor='k', label=r'$\tau$ = 16.5', normed=1)
ax2.hist(hiss[1][0], x, color='g', edgecolor='k', label=r'$\tau$ = 22', normed=1)
ax3.hist(hiss[2][0], x, color='lightcoral', edgecolor='k', label=r'$\tau$ = 29', normed=1)
ax1.plot(x, y1, 'r--')
ax2.plot(x, y2, 'r--')
ax3.plot(x, y3, 'r--')
ax3.set_xlabel(r'$\psi$ (h)',fontsize=20, labelpad=20)
for a in f.axes: a.set_ylabel('probability',fontsize=20, labelpad=20); a.legend(fontsize=16); a.set_yticks(np.arange(0, 0.3, 0.1)); a.set_xticks(np.arange(x[0], x[-1], 2));
ax1.tick_params(labelbottom='off')
ax2.tick_params(labelbottom='off')
f.subplots_adjust(hspace=0)
plt.show()
#%%ttest amplitude ist abhängig von tau
print('16.5 vs 22.:', scipy.stats.ttest_ind(hiss[0][1], hiss[1][1]))
print('16.5 vs 29.:', scipy.stats.ttest_ind(hiss[0][1], hiss[2][1]))
print('22   vs 29.:', scipy.stats.ttest_ind(hiss[1][1], hiss[2][1]))
#%%histogram alle einzeln
hiss = []
bins = [0,0]
for j in range(3):
    his = []
    for i in range(9):
        psis = np.array([array[5][j*27+i], array[5][j*27+i+9], array[5][j*27+i+18]])
        for l in range(3):
            if psis[l]>12: psis[l]=psis[l]-24
        entrained = np.array([array[6][j*27+i], array[6][j*27+i+9], array[6][j*27+i+18]])
        psis = psis[entrained<2]
        his.append(psis)
        if max(psis)>bins[1]: bins[1] = max(psis)
        if min(psis)<bins[0]: bins[0] = min(psis)
    hiss.append(his)
    
x = np.arange(round(bins[0]), round(bins[1]), 1)

for i in range(3):
    f, ((ax1), (ax2), (ax3), (ax4), (ax5), (ax6), (ax7), (ax8), (ax9)) = plt.subplots(9,1, figsize=(12,12))
    axes = [(ax1), (ax2), (ax3), (ax4), (ax5), (ax6), (ax7), (ax8), (ax9)]
    for k, v in enumerate(axes):
        y = mlab.normpdf(x, np.mean(hiss[i][k]), np.std(hiss[i][k]))
        v.hist(hiss[i][k], x, edgecolor='k', label='tau = 16.5', normed=1)
        v.plot(x, y, 'r--')
        #v.legend()
        v.set_yticks(np.arange(0, 1.1, 1))
        v.set_xticks(np.arange(round(bins[0]), round(bins[1]), 2))
        v.tick_params(labelbottom='off') 
    ax9.set_xlabel(r'$\psi$ (h)',fontsize=14)
    ax9.tick_params(labelbottom='on') 
    ax5.set_ylabel(r'$probability$',fontsize=14)
    f.suptitle('tau = '+ str(array[0][i*27]),fontsize=14)
    f.subplots_adjust(top=0.95)
"""